import io
import calendar
from datetime import date, datetime
from typing import TYPE_CHECKING

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if TYPE_CHECKING:
    from .models import Funcionario
    from .time_service import ResultadoDia, ResumoMes


STATUS_PT = {
    "completo": "Completo",
    "atrasado": "Atrasado",
    "saida_antecipada": "Saída antecipada",
    "incompleto": "Incompleto",
    "falta": "Falta",
    "falta_justificada": "Falta justificada",
    "feriado": "Feriado",
    "folga": "Folga",
}

DIAS_PT = ["", "Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


def fmt_min(m: int) -> str:
    sinal = "-" if m < 0 else ""
    m = abs(m)
    return f"{sinal}{m // 60:02d}h{m % 60:02d}min"


def gerar_pdf_espelho(
    funcionario: "Funcionario",
    dias: list,
    resumo: "ResumoMes",
    mes: int,
    ano: int,
    nome_empresa: str = "",
) -> bytes:
    """Gera PDF do espelho de ponto individual. Retorna bytes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=styles["Heading1"],
                                   alignment=TA_CENTER, fontSize=14)
    estilo_normal = styles["Normal"]

    meses_pt = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

    elementos = []

    if nome_empresa:
        elementos.append(Paragraph(nome_empresa, estilo_titulo))
    elementos.append(Paragraph("ESPELHO DE PONTO", estilo_titulo))
    elementos.append(Spacer(1, 0.3*cm))

    info = [
        ["Funcionário:", funcionario.nome, "Matrícula:", funcionario.matricula or ""],
        ["CPF:", funcionario.cpf or "", "Cargo:", funcionario.cargo or ""],
        ["Departamento:", funcionario.departamento or "", "Período:", f"{meses_pt[mes]}/{ano}"],
    ]
    tabela_info = Table(info, colWidths=[3.5*cm, 6*cm, 3*cm, 4.5*cm])
    tabela_info.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabela_info)
    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elementos.append(Spacer(1, 0.3*cm))

    cabecalho = ["Data", "Dia", "Entrada", "Saída", "Horas", "Saldo", "Status"]
    dados_tabela = [cabecalho]

    for r in dias:
        entrada_str = r.entrada.strftime("%H:%M") if r.entrada else "-"
        saida_str = r.saida.strftime("%H:%M") if r.saida else "-"
        linha = [
            r.data.strftime("%d/%m/%Y"),
            DIAS_PT[r.data.isoweekday()],
            entrada_str,
            saida_str,
            fmt_min(r.horas_trabalhadas_min),
            fmt_min(r.saldo_min),
            STATUS_PT.get(r.status.value, r.status.value),
        ]
        dados_tabela.append(linha)

    tabela_pontos = Table(dados_tabela,
        colWidths=[2.5*cm, 1.2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 3.3*cm])
    estilo_tabela = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ])
    for i, r in enumerate(dias, start=1):
        if r.status.value == "falta":
            estilo_tabela.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#fee2e2"))
        elif r.status.value in ("feriado", "folga"):
            estilo_tabela.add("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#94a3b8"))
    tabela_pontos.setStyle(estilo_tabela)
    elementos.append(tabela_pontos)
    elementos.append(Spacer(1, 0.5*cm))

    totais = [
        ["Horas trabalhadas:", fmt_min(resumo.total_trabalhado_min),
         "Horas previstas:", fmt_min(resumo.total_previsto_min)],
        ["Saldo do mês:", fmt_min(resumo.saldo_min), "Faltas:", str(resumo.faltas)],
        ["Atrasos:", str(resumo.atrasos), "Faltas justificadas:", str(resumo.faltas_justificadas)],
    ]
    tabela_totais = Table(totais, colWidths=[5*cm, 3.5*cm, 5*cm, 3.5*cm])
    tabela_totais.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f1f5f9")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elementos.append(tabela_totais)
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph("* Batida editada manualmente pelo administrador.", estilo_normal))
    elementos.append(Spacer(1, 1.5*cm))

    assinaturas = [
        [f"_________________________\n{funcionario.nome}\nFuncionário",
         "_________________________\nResponsável RH"],
    ]
    tabela_ass = Table(assinaturas, colWidths=[8*cm, 8*cm])
    tabela_ass.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabela_ass)

    doc.build(elementos)
    return buffer.getvalue()


def gerar_excel_consolidado(
    funcionarios_dados: list,
    mes: int,
    ano: int,
    nome_empresa: str = "",
    feriados: list = None,
) -> bytes:
    """Gera Excel consolidado mensal. Retorna bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    meses_pt = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    ws.title = f"{meses_pt[mes]} {ano}"

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    total_fill = PatternFill("solid", fgColor="DBEAFE")
    alt_fill = PatternFill("solid", fgColor="F1F5F9")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Relatório Consolidado — {meses_pt[mes]}/{ano}" + (f" — {nome_empresa}" if nome_empresa else "")
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    colunas = ["Matrícula", "Nome", "Departamento", "Dias Trabalhados",
               "Horas Trabalhadas", "Horas Previstas", "Saldo",
               "Atrasos (qtd)", "Faltas (qtd)", "Faltas Justif. (qtd)"]
    for col_idx, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=2, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20

    for row_idx, d in enumerate(funcionarios_dados, start=3):
        valores = [
            d["matricula"], d["nome"], d["departamento"], d["dias_trabalhados"],
            fmt_min(d["total_trabalhado_min"]), fmt_min(d["total_previsto_min"]),
            fmt_min(d["saldo_min"]), d["atrasos"], d["faltas"], d["faltas_justificadas"],
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = normal_font
            cell.alignment = center if col_idx != 2 else Alignment(horizontal="left", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill

    total_row = len(funcionarios_dados) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = bold_font
    ws.cell(row=total_row, column=4, value=sum(d["dias_trabalhados"] for d in funcionarios_dados)).font = bold_font
    ws.cell(row=total_row, column=8, value=sum(d["atrasos"] for d in funcionarios_dados)).font = bold_font
    ws.cell(row=total_row, column=9, value=sum(d["faltas"] for d in funcionarios_dados)).font = bold_font
    ws.cell(row=total_row, column=10, value=sum(d["faltas_justificadas"] for d in funcionarios_dados)).font = bold_font
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).alignment = center

    larguras = [12, 30, 20, 16, 18, 16, 14, 14, 14, 18]
    for col_idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = largura

    ws.freeze_panes = "A3"

    if feriados:
        ws2 = wb.create_sheet(title="Feriados")
        ws2["A1"] = "Data"
        ws2["B1"] = "Descrição"
        ws2["C1"] = "Recorrente"
        for i, f in enumerate(feriados, start=2):
            ws2.cell(row=i, column=1, value=str(f.data))
            ws2.cell(row=i, column=2, value=f.descricao)
            ws2.cell(row=i, column=3, value="Sim" if f.recorrente_anual else "Não")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
